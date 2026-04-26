"""
One-shot script: lee hierro.xlsx y emite tests/fixtures/ejemplo_real.json
con elementos por (hoja, phi) y los targets optimizados manuales.

Re-correr solo si cambia el Excel de referencia.

Uso:
    .venv/Scripts/python scripts/extract_fixtures.py
"""

from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "hierro.xlsx"
OUT = ROOT / "tests" / "fixtures" / "ejemplo_real.json"

# Hojas a procesar y nombres "slug" para el JSON
SHEETS = {
    "bases y vf": "bases_y_vf",
    "tronco y tabique": "tronco_y_tabique",
    # "6° a 9°" lo dejamos afuera por ahora — la usuaria no lo flagueó como caso de test
}


def parse_sheet(ws) -> tuple[dict[int, list[dict]], dict[int, int], dict[int, int]]:
    """
    Recorre una hoja y devuelve:
      - elementos_por_phi: {phi: [{nombre, phi, cant_elementos, cant_rep, medida}, ...]}
      - target_optimizado: {phi: barras_optimizadas} (de las filas donde col2=phi y col13=int)
      - target_ingenuo:    {phi: barras_metodo_ingenuo} (de las filas de subtotal)
    """
    elementos: dict[int, list[dict]] = defaultdict(list)
    optimizado: dict[int, int] = {}
    ingenuo: dict[int, int] = {}

    current_phi: int | None = None
    pending_ingenuo: int | None = None  # último subtotal "ingenuo" visto, asociado al próximo φ-row

    for r in range(1, ws.max_row + 1):
        c1 = ws.cell(row=r, column=1).value   # nombre
        c2 = ws.cell(row=r, column=2).value   # phi
        c3 = ws.cell(row=r, column=3).value   # cant elementos
        c4 = ws.cell(row=r, column=4).value   # cant repeticiones
        c6 = ws.cell(row=r, column=6).value   # medida
        c13 = ws.cell(row=r, column=13).value  # totales (col M)

        # Fila de elemento: nombre + phi + cant + medida
        es_elemento = (
            isinstance(c1, str)
            and c1.strip()
            and isinstance(c2, int)
            and isinstance(c3, (int, float))
            and isinstance(c6, (int, float))
        )

        if es_elemento:
            phi = int(c2)
            current_phi = phi
            elementos[phi].append({
                "nombre": c1.strip(),
                "phi": phi,
                "cant_elementos": int(c3),
                "cant_rep": int(c4) if isinstance(c4, (int, float)) else 1,
                "medida": float(c6),
            })
            continue

        # Fila de subtotal "ingenuo": col1 vacía, col2 vacía, col13 entero (suma de cantidad final)
        if c1 is None and c2 is None and isinstance(c13, (int, float)) and c13 > 0:
            pending_ingenuo = int(c13)
            continue

        # Fila de total "optimizado": col1 vacía, col2 = phi (entero), col13 entero
        if c1 is None and isinstance(c2, int) and isinstance(c13, (int, float)) and c13 > 0:
            phi = int(c2)
            optimizado[phi] = int(c13)
            if pending_ingenuo is not None:
                ingenuo[phi] = pending_ingenuo
                pending_ingenuo = None

    return dict(elementos), optimizado, ingenuo


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    out = {
        "source": "hierro.xlsx",
        "largo_barra": 12.0,
        "data": {},
        "expected_targets": {},
        "ingenuo_referencia": {},
    }

    for sheet_name, slug in SHEETS.items():
        ws = wb[sheet_name]
        elementos, optimizado, ingenuo = parse_sheet(ws)
        out["data"][slug] = {str(phi): elems for phi, elems in elementos.items()}
        for phi, target in optimizado.items():
            out["expected_targets"][f"{slug}:{phi}"] = target
        for phi, val in ingenuo.items():
            out["ingenuo_referencia"][f"{slug}:{phi}"] = val

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f"OK -> {OUT}")
    print("Targets optimizados:")
    for k, v in out["expected_targets"].items():
        ing = out["ingenuo_referencia"].get(k, "?")
        print(f"  {k:30s} ingenuo={ing}  ->  optimizado<={v}")


if __name__ == "__main__":
    main()
