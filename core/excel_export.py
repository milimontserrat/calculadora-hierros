"""
Generación del Excel descargable.

Replica la estructura de la planilla original (`hierro.xlsx`) con todas las
columnas calculadas mediante fórmulas de Excel (cantidad total, longitud total,
cantidad directa, cantidad x barra, cantidad final de barras, sobrante).

Agrega dos hojas adicionales:
  - Resumen: comparación φ por φ entre método ingenuo y optimizado.
  - Plan de corte: cada barra física con qué piezas se cortan de ella.
"""
from __future__ import annotations

from collections import Counter, defaultdict
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core.models import CutPlan, Element

# Paleta consistente con la app
_COLOR_HEADER = "C2185B"
_COLOR_HEADER_TEXT = "FFFFFF"
_COLOR_OPT = "C2185B"
_COLOR_SUBTOTAL_BG = "FDF2F8"


def generar_excel(
    nombre_proyecto: str,
    elementos: list[Element],
    plan: CutPlan,
    largo_barra: float,
    ingenuo_por_phi: dict[int, int],
) -> bytes:
    """Devuelve los bytes del .xlsx listo para descargar."""
    wb = Workbook()
    ws_planilla = wb.active
    ws_planilla.title = "Planilla"

    _escribir_planilla(ws_planilla, nombre_proyecto, elementos, largo_barra, plan, ingenuo_por_phi)
    _escribir_resumen(wb, plan, ingenuo_por_phi, largo_barra)
    _escribir_plan_corte(wb, plan)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# -------------------- Hoja Planilla --------------------

_HEADERS_PLANILLA = [
    "Elemento",                  # A
    "φ",                         # B
    "Cant. elementos",           # C
    "Cant. repeticiones",        # D
    "Cant. total",               # E  =C*D
    "Medida (m)",                # F
    "Longitud total (m)",        # G  =E*F
    "Cant. directa",             # H  =G/L
    "Cant. por barra",           # I  =FLOOR(L/F,1)
    "Cant. de barras",           # J  =E/I  (sin redondear)
    "Cant. final de barras",     # K  =CEILING(E/I,1)
    "Sobrante por barra (m)",    # L  =L_largo - I*F
]


def _aplicar_estilo_header(ws: Worksheet, fila: int, n_cols: int, alto: int = 36) -> None:
    fill = PatternFill("solid", fgColor=_COLOR_HEADER)
    font = Font(bold=True, color=_COLOR_HEADER_TEXT, size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, n_cols + 1):
        c = ws.cell(row=fila, column=col)
        c.fill = fill
        c.font = font
        c.alignment = align
    ws.row_dimensions[fila].height = alto


def _escribir_planilla(
    ws: Worksheet,
    nombre_proyecto: str,
    elementos: list[Element],
    largo_barra: float,
    plan: CutPlan,
    ingenuo_por_phi: dict[int, int],
) -> None:
    # Título
    ws.cell(row=1, column=1, value=nombre_proyecto).font = Font(bold=True, size=14, color=_COLOR_HEADER)
    ws.cell(row=2, column=1, value=f"Largo de barra: {largo_barra} m").font = Font(italic=True, size=10)

    # Headers en fila 4
    for col_idx, h in enumerate(_HEADERS_PLANILLA, start=1):
        ws.cell(row=4, column=col_idx, value=h)
    _aplicar_estilo_header(ws, 4, len(_HEADERS_PLANILLA))

    # Agrupar por phi
    elems_por_phi: dict[int, list[Element]] = defaultdict(list)
    for e in elementos:
        if e.es_valido():
            elems_por_phi[e.phi].append(e)

    fill_subtotal = PatternFill("solid", fgColor=_COLOR_SUBTOTAL_BG)
    row = 5

    for phi in sorted(elems_por_phi.keys()):
        for e in elems_por_phi[phi]:
            r = row
            ws.cell(row=r, column=1, value=e.nombre)
            ws.cell(row=r, column=2, value=e.phi)
            ws.cell(row=r, column=3, value=e.cantidad_elementos)
            ws.cell(row=r, column=4, value=e.cantidad_repeticiones)
            ws.cell(row=r, column=5, value=f"=C{r}*D{r}")
            ws.cell(row=r, column=6, value=e.medida)
            ws.cell(row=r, column=7, value=f"=E{r}*F{r}")
            ws.cell(row=r, column=8, value=f"=G{r}/{largo_barra}")
            ws.cell(row=r, column=9, value=f"=FLOOR({largo_barra}/F{r},1)")
            ws.cell(row=r, column=10, value=f"=IF(I{r}=0,\"\",E{r}/I{r})")  # cantidad sin redondear
            ws.cell(row=r, column=11, value=f"=IF(I{r}=0,\"\",CEILING(E{r}/I{r},1))")
            ws.cell(row=r, column=12, value=f"={largo_barra}-I{r}*F{r}")
            # Formato decimales
            for col in (6, 7, 8, 10, 12):
                ws.cell(row=r, column=col).number_format = "0.00"
            row += 1

        # Subtotal por phi: ingenuo vs optimizado (en columna K = "Cant. final de barras")
        n_opt = len(plan.barras_por_phi.get(phi, []))
        n_ing = ingenuo_por_phi.get(phi, n_opt)

        ws.cell(row=row, column=1, value=f"Subtotal φ{phi} — método manual").font = Font(italic=True)
        ws.cell(row=row, column=11, value=n_ing).font = Font(bold=True)
        for col in range(1, 13):
            ws.cell(row=row, column=col).fill = fill_subtotal
        row += 1

        ws.cell(row=row, column=1, value=f"Subtotal φ{phi} — OPTIMIZADO").font = Font(italic=True, bold=True, color=_COLOR_OPT)
        ws.cell(row=row, column=11, value=n_opt).font = Font(bold=True, color=_COLOR_OPT)
        for col in range(1, 13):
            ws.cell(row=row, column=col).fill = fill_subtotal
        row += 2  # línea en blanco entre φ

    # Anchos de columna (12 columnas)
    widths = [24, 6, 16, 18, 12, 12, 18, 14, 16, 16, 20, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"


# -------------------- Hoja Resumen --------------------

def _escribir_resumen(
    wb: Workbook,
    plan: CutPlan,
    ingenuo_por_phi: dict[int, int],
    largo_barra: float,
) -> None:
    ws = wb.create_sheet("Resumen")
    headers = ["φ", "Barras (manual)", "Barras (optimizado)",
               "Ahorro (barras)", "Ahorro (m)", "Desperdicio (m)"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    _aplicar_estilo_header(ws, 1, len(headers))

    row = 2
    total_ing = total_opt = 0
    total_desp = 0.0
    for phi in sorted(plan.barras_por_phi.keys()):
        barras = plan.barras_por_phi[phi]
        n_opt = len(barras)
        n_ing = ingenuo_por_phi.get(phi, n_opt)
        ahorro = n_ing - n_opt
        desp = sum(b.sobrante for b in barras)

        ws.cell(row=row, column=1, value=phi)
        ws.cell(row=row, column=2, value=n_ing)
        ws.cell(row=row, column=3, value=n_opt)
        ws.cell(row=row, column=4, value=ahorro)
        ws.cell(row=row, column=5, value=round(ahorro * largo_barra, 2))
        ws.cell(row=row, column=6, value=round(desp, 2))
        for col in (5, 6):
            ws.cell(row=row, column=col).number_format = "0.00"

        total_ing += n_ing
        total_opt += n_opt
        total_desp += desp
        row += 1

    # Fila TOTAL
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=2, value=total_ing)
    ws.cell(row=row, column=3, value=total_opt)
    ws.cell(row=row, column=4, value=total_ing - total_opt)
    ws.cell(row=row, column=5, value=round((total_ing - total_opt) * largo_barra, 2))
    ws.cell(row=row, column=6, value=round(total_desp, 2))
    bold = Font(bold=True, size=11, color=_COLOR_OPT)
    fill = PatternFill("solid", fgColor=_COLOR_SUBTOTAL_BG)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=row, column=col)
        c.font = bold
        c.fill = fill
    for col in (5, 6):
        ws.cell(row=row, column=col).number_format = "0.00"

    widths = [8, 18, 22, 16, 14, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# -------------------- Hoja Plan de corte --------------------

def _formato_piezas(piezas: list[tuple[str, float]]) -> str:
    medidas = [round(m, 2) for _, m in piezas]
    cnt = Counter(medidas)
    return " + ".join(f"{n}×{m:.2f}" for m, n in sorted(cnt.items(), reverse=True))


def _escribir_plan_corte(wb: Workbook, plan: CutPlan) -> None:
    ws = wb.create_sheet("Plan de corte")
    headers = ["φ", "Barra #", "Piezas cortadas", "Sobrante (m)"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    _aplicar_estilo_header(ws, 1, len(headers))

    row = 2
    for phi in sorted(plan.barras_por_phi.keys()):
        for idx, b in enumerate(plan.barras_por_phi[phi], start=1):
            ws.cell(row=row, column=1, value=phi)
            ws.cell(row=row, column=2, value=idx)
            ws.cell(row=row, column=3, value=_formato_piezas(b.piezas))
            ws.cell(row=row, column=4, value=round(b.sobrante, 2))
            ws.cell(row=row, column=4).number_format = "0.00"
            row += 1

    widths = [8, 10, 70, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
