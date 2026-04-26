"""
Parser tolerante de planillas de armadura en Excel.

Detecta automáticamente las columnas relevantes por nombre (case-insensitive,
ignorando tildes y espacios). Recorre todas las hojas. Ignora filas vacías
y filas de subtotal (las que tienen NaN en las columnas críticas pero algún
número en columnas agregadas).

Uso:
    from core.excel_import import parse_excel
    elementos = parse_excel(uploaded_file.getvalue())  # bytes
"""
from __future__ import annotations

import io
import unicodedata
from collections.abc import Iterable

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from core.models import Element

# Sinónimos por columna lógica. Cada entry es la lista de variantes aceptadas
# para esa columna (después de normalizar: minúsculas, sin tildes, sin espacios).
_SINONIMOS: dict[str, list[str]] = {
    "nombre": ["elemento", "nombre", "descripcion", "item"],
    "phi": ["fi", "phi", "diametro", "diam", "φ"],
    "cant_elementos": [
        "cantidaddeelementos", "cantelementos", "cantidad",
        "cant", "cantelem", "cantidadelementos",
    ],
    "cant_rep": [
        "cantidadderepeticiones", "repeticiones", "rep",
        "cantrep", "cantidadrepeticiones", "repet",
    ],
    "medida": ["medida", "largo", "longitud", "largom", "medidam"],
}


def _normalizar(s: object) -> str:
    """Pasa a minúsculas, saca tildes, espacios y caracteres no alfanuméricos."""
    if s is None:
        return ""
    txt = str(s).strip().lower()
    nfkd = unicodedata.normalize("NFKD", txt)
    txt = "".join(c for c in nfkd if not unicodedata.combining(c))
    return "".join(c for c in txt if c.isalnum())


def _detectar_columnas(ws: Worksheet, max_filas_busqueda: int = 10) -> tuple[dict[str, int], int] | None:
    """
    Recorre las primeras filas buscando headers. Devuelve ({logical -> col_idx}, header_row)
    o None si ninguna fila tiene la estructura mínima.

    Estrategia:
      - Pass 1: match exacto (preferido).
      - Pass 2: match por substring (solo sobre columnas no asignadas todavía).
      - Una columna puede asignarse a un único logical (no se reutiliza).
      - Si no se detecta "nombre" pero el resto sí, asume que es la columna 1
        (caso común en el Excel original sin header en col A).
    """
    for r in range(1, min(max_filas_busqueda + 1, ws.max_row + 1)):
        headers = {c: _normalizar(ws.cell(row=r, column=c).value)
                   for c in range(1, ws.max_column + 1)}
        candidatos: dict[str, int] = {}
        usadas: set[int] = set()

        # Pass 1: match exacto
        for logical, syns in _SINONIMOS.items():
            for c, h in headers.items():
                if c in usadas or not h:
                    continue
                if h in syns:
                    candidatos[logical] = c
                    usadas.add(c)
                    break

        # Pass 2: contains, solo sobre columnas libres
        for logical, syns in _SINONIMOS.items():
            if logical in candidatos:
                continue
            for c, h in headers.items():
                if c in usadas or not h:
                    continue
                if any(s in h for s in syns):
                    candidatos[logical] = c
                    usadas.add(c)
                    break

        # Fallback para "nombre": si están phi+cant+medida pero no nombre, asumir col 1
        if (
            "nombre" not in candidatos
            and all(k in candidatos for k in ("phi", "cant_elementos", "medida"))
            and 1 not in usadas
        ):
            candidatos["nombre"] = 1

        # Mínimo necesario
        if all(k in candidatos for k in ("nombre", "phi", "cant_elementos", "medida")):
            return candidatos, r

    return None


def _to_int(v: object) -> int | None:
    if v is None:
        return None
    try:
        f = float(v)
        if f != f:  # NaN
            return None
        return int(round(f))
    except (TypeError, ValueError):
        return None


def _to_float(v: object) -> float | None:
    if v is None:
        return None
    try:
        f = float(v)
        if f != f:  # NaN
            return None
        return f
    except (TypeError, ValueError):
        return None


def _parse_hoja(ws: Worksheet) -> list[Element]:
    """Devuelve los elementos parseados de una hoja, o lista vacía si no se detectan headers."""
    detectado = _detectar_columnas(ws)
    if detectado is None:
        return []
    cols, header_row = detectado

    elementos: list[Element] = []
    for r in range(header_row + 1, ws.max_row + 1):
        nombre_raw = ws.cell(row=r, column=cols["nombre"]).value
        phi_raw = ws.cell(row=r, column=cols["phi"]).value
        cant_raw = ws.cell(row=r, column=cols["cant_elementos"]).value
        medida_raw = ws.cell(row=r, column=cols["medida"]).value
        rep_raw = ws.cell(row=r, column=cols["cant_rep"]).value if "cant_rep" in cols else 1

        nombre = str(nombre_raw).strip() if nombre_raw is not None else ""
        phi = _to_int(phi_raw)
        cant = _to_int(cant_raw)
        medida = _to_float(medida_raw)
        rep = _to_int(rep_raw) or 1

        # Filtros para descartar filas no-elemento (vacías, subtotales, separadores)
        if not nombre:
            continue
        if phi is None or phi <= 0:
            continue
        if cant is None or cant <= 0:
            continue
        if medida is None or medida <= 0:
            continue

        elementos.append(Element(
            nombre=nombre,
            phi=phi,
            cantidad_elementos=cant,
            cantidad_repeticiones=rep,
            medida=medida,
        ))

    return elementos


def parse_excel(file_bytes: bytes) -> list[Element]:
    """
    Parsea un .xlsx subido por la usuaria. Recorre TODAS las hojas y concatena
    los elementos detectados. Levanta ValueError si no encuentra ninguna hoja
    con headers válidos.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(f"No pude abrir el archivo: {e}") from e

    todos: list[Element] = []
    hojas_validas: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        elems = _parse_hoja(ws)
        if elems:
            hojas_validas.append(sheet_name)
            todos.extend(elems)

    if not todos:
        raise ValueError(
            "No detecté ninguna hoja con la estructura esperada. "
            "El archivo debe tener columnas que se llamen aproximadamente: "
            "Elemento, φ (o fi/diámetro), Cantidad de elementos, Medida."
        )

    return todos
